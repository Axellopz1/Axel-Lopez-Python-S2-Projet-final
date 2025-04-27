# --- Imports ---
import pandas as pd  # Manipulation de DataFrame pour organiser les données
import matplotlib.pyplot as plt  # Pour créer les graphiques
from openpyxl import load_workbook  # Pour manipuler un fichier Excel existant
from openpyxl.styles import PatternFill, Font  # Pour styliser les cellules Excel (couleurs, polices)
from openpyxl.utils.dataframe import dataframe_to_rows  # (pas utilisé ici mais sert à convertir DataFrame en lignes Excel)
from openpyxl.drawing.image import Image as XLImage  # Pour insérer des images (graphiques) dans Excel
import io  # Manipulation de flux d'images en mémoire

from config import Config  # Pour utiliser le chemin d'export défini dans config.py

# --- Fonction principale pour exporter tout vers un fichier Excel ---
def export_vers_excel(portefeuille, projections, stress_results):
    # Crée un nouveau fichier Excel avec Openpyxl
    writer = pd.ExcelWriter(Config.FICHIER_EXPORT, engine='openpyxl')

    # --- Feuille 1 : Portefeuille ---
    df_portefeuille = portefeuille.composition()  # Récupère la composition actuelle du portefeuille
    valeur_totale = df_portefeuille["Valeur Totale"].sum()  # Calcule la valeur totale du portefeuille

    # Ajoute une ligne "TOTAL" à la fin du tableau
    total_values = ["TOTAL"] + [""] * (len(df_portefeuille.columns) - 2) + [valeur_totale]
    total_row = pd.DataFrame([total_values], columns=df_portefeuille.columns)
    df_final = pd.concat([df_portefeuille, total_row])

    # Écrit cette feuille dans Excel
    df_final.to_excel(writer, sheet_name='Portefeuille', index=False)

    # --- Feuille 2 : Projections ---
    projections_copy = projections.copy()  # Copie les projections pour ne pas modifier l'original

    # Pour chaque scénario, calcule la performance en pourcentage
    for scenario in projections_copy.columns:
        projections_copy[scenario + " Gain/Perte (%)"] = (projections_copy[scenario] / projections_copy[scenario].iloc[0] - 1) * 100

    # Écrit les projections et performances dans une nouvelle feuille
    projections_copy.to_excel(writer, sheet_name='Projections')

    # --- Feuille 3 : Stress Tests ---
    stress_results_copy = stress_results.copy()  # Copie les stress tests
    pertes_dollars = {}  # Dictionnaire pour stocker les pertes en dollars
    pertes_pourcentage = {}  # Dictionnaire pour stocker les pertes en pourcentage

    # Pour chaque scénario de stress
    for col in stress_results.columns:
        valeur_init = stress_results[col].iloc[0]
        valeur_fin = stress_results[col].iloc[-1]
        pertes_dollars[col] = valeur_init - valeur_fin
        pertes_pourcentage[col] = ((valeur_init - valeur_fin) / valeur_init) * 100

    # Ajoute les lignes Perte ($) et Perte (%) dans le DataFrame
    stress_results_copy.loc["Perte ($)"] = pertes_dollars
    stress_results_copy.loc["Perte (%)"] = pertes_pourcentage

    # Écrit les résultats des stress tests dans Excel
    stress_results_copy.to_excel(writer, sheet_name='Stress Tests')

    # --- Feuille 4 : Analyse Quantitative ---
    analyse = analyse_quantitative(projections)  # Calcul des indicateurs de performance
    analyse.to_excel(writer, sheet_name='Analyse Quantitative')

    writer.close()  # Ferme et sauvegarde le fichier Excel

    # --- Post-traitement : Ajout du style et des graphiques ---
    workbook = load_workbook(Config.FICHIER_EXPORT)  # Recharge le fichier Excel pour le modifier

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Mettre en bleu clair l'entête de la première ligne
        for cell in sheet[1]:
            cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            cell.font = Font(bold=True)

        # Ajuste automatiquement la largeur de chaque colonne selon le contenu
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        # --- Générer un graphique pour chaque feuille ---
        if sheet_name == 'Portefeuille':
            plot = portefeuille_plot(df_portefeuille)
        elif sheet_name == 'Projections':
            plot = projection_plot(projections)
        elif sheet_name == 'Stress Tests':
            plot = stress_plot(stress_results)
        elif sheet_name == 'Analyse Quantitative':
            plot = None  # Pas de graphique pour l'analyse quantitative

        # Si un graphique a été généré, l'insérer dans la feuille
        if plot:
            img_bytes = io.BytesIO()
            plot.savefig(img_bytes, format='png')
            img_bytes.seek(0)
            img = XLImage(img_bytes)
            img.width = 600
            img.height = 400
            sheet.add_image(img, "H2")  # Insère l'image dans la cellule H2

    workbook.save(Config.FICHIER_EXPORT)  # Sauvegarde finale du fichier Excel

# --- Fonctions auxiliaires pour générer les graphiques ---

def portefeuille_plot(df):
    fig, ax = plt.subplots()
    df.groupby("Type")["Valeur Totale"].sum().plot(kind="pie", autopct='%1.1f%%', ax=ax)
    ax.set_ylabel("")
    ax.set_title("Répartition du Portefeuille")
    return fig

def projection_plot(projections):
    fig, ax = plt.subplots()
    projections.plot(ax=ax)
    ax.set_title("Projection de Valeur du Portefeuille (24 mois)")
    return fig

def stress_plot(stress_results):
    fig, ax = plt.subplots()
    stress_results.drop(["Perte ($)", "Perte (%)"], axis=0, errors='ignore').plot(ax=ax)
    ax.set_title("Stress Tests du Portefeuille")
    return fig

# --- Fonction pour l'analyse quantitative du portefeuille ---

def analyse_quantitative(projections):
    valeur_initiale = projections.iloc[0]  # Prend la première valeur
    valeur_finale = projections.iloc[-1]  # Prend la dernière valeur
    rendement_total = ((valeur_finale - valeur_initiale) / valeur_initiale) * 100  # Gain/perte total en %
    cagr = ((valeur_finale / valeur_initiale) ** (1/2) - 1) * 100  # Rendement annualisé (sur 2 ans ici)
    volatilite_mensuelle = projections.pct_change().std() * 100  # Volatilité mensuelle moyenne
    max_drawdown = (projections.cummax() - projections).max() / projections.cummax().max() * 100  # Plus forte perte maximale

    # Résultat sous forme de DataFrame pour l'écrire dans Excel
    analyse = pd.DataFrame({
        "Valeur Initiale ($)": valeur_initiale,
        "Valeur Finale ($)": valeur_finale,
        "Rendement Total (%)": rendement_total,
        "CAGR (%)": cagr,
        "Volatilité Mensuelle (%)": volatilite_mensuelle,
        "Maximum Drawdown (%)": max_drawdown
    })

    return analyse
